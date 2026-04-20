from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from typing import Optional
from io import BytesIO

from database import get_db
from models import HighwayAsset, Measurement, MaintenanceOrder
from schemas import ComplianceReport

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    _HAS_RL = True
except Exception:
    _HAS_RL = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False


router = APIRouter(prefix="/api/reports", tags=["Reports"])


# ── JSON compliance report (used by frontend) ──────────────────────────────

@router.get("/compliance", response_model=ComplianceReport)
def compliance_report(
    highway_id: str = Query(..., description="Highway ID, e.g. NH-48"),
    db: Session = Depends(get_db),
):
    assets = db.query(HighwayAsset).filter(HighwayAsset.highway_id == highway_id).all()
    if not assets:
        raise HTTPException(404, f"No assets found for {highway_id}")

    total = len(assets)
    compliant = sum(1 for a in assets if a.status == "compliant")
    warning = sum(1 for a in assets if a.status == "warning")
    critical = sum(1 for a in assets if a.status == "critical")

    by_type: dict = {}
    for a in assets:
        t = a.asset_type
        if t not in by_type:
            by_type[t] = {"total": 0, "compliant": 0, "warning": 0, "critical": 0}
        by_type[t]["total"] += 1
        by_type[t][a.status] += 1

    critical_assets = [a for a in assets if a.status == "critical"]

    return ComplianceReport(
        highway_id=highway_id,
        generated_at=datetime.utcnow(),
        total_assets=total,
        compliant=compliant,
        warning=warning,
        critical=critical,
        compliance_pct=round((compliant / total) * 100, 1) if total else 0.0,
        assets_by_type=by_type,
        critical_assets=critical_assets,
    )


# ── PDF reports ─────────────────────────────────────────────────────────────

def _require_rl():
    if not _HAS_RL:
        raise HTTPException(
            503, "reportlab not installed. `pip install reportlab` to enable PDFs."
        )


def _compliance_pdf(highway_id: str, assets, standard: str) -> bytes:
    """Render a compliance PDF for IRC 67 (signs) or IRC 35 (markings)."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=colors.HexColor("#0f3057"))
    flow = []

    flow.append(Paragraph(f"{standard} Compliance Report", h1))
    flow.append(Paragraph(f"Highway: <b>{highway_id}</b>", styles["Normal"]))
    flow.append(Paragraph(
        f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}", styles["Normal"]
    ))
    flow.append(Spacer(1, 12))

    total = len(assets)
    compliant = sum(1 for a in assets if a.status == "compliant")
    warning = sum(1 for a in assets if a.status == "warning")
    critical = sum(1 for a in assets if a.status == "critical")
    pct = round((compliant / total) * 100, 1) if total else 0.0

    summary = [
        ["Total assets", total],
        ["Compliant", f"{compliant} ({pct}%)"],
        ["Warning", warning],
        ["Critical", critical],
    ]
    t = Table(summary, colWidths=[6 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2f7")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    flow.append(t)
    flow.append(Spacer(1, 16))

    flow.append(Paragraph("Critical / Non-compliant assets", styles["Heading2"]))
    head = ["ID", "Type", "Chainage km", "Current R_L", "IRC min", "Status"]
    rows = [head]
    for a in sorted(assets, key=lambda x: (x.status != "critical", x.chainage_km)):
        rows.append([
            a.id,
            a.asset_type,
            round(a.chainage_km, 2),
            round(a.current_rl, 1) if a.current_rl else "—",
            round(a.irc_minimum_rl, 1),
            a.status,
        ])
        if len(rows) > 80:
            break

    body = Table(rows, repeatRows=1)
    body.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 1), (4, -1), "RIGHT"),
    ]))
    flow.append(body)

    doc.build(flow)
    return buf.getvalue()


@router.get("/compliance/pdf")
def compliance_pdf(
    highway_id: str = Query(...),
    standard: str = Query("IRC-67", pattern="^(IRC-67|IRC-35)$"),
    db: Session = Depends(get_db),
):
    _require_rl()
    q = db.query(HighwayAsset).filter(HighwayAsset.highway_id == highway_id)
    if standard == "IRC-67":
        q = q.filter(HighwayAsset.asset_type.in_(["sign", "regulatory_sign", "warning_sign", "informatory_sign", "gantry_sign"]))
    else:
        q = q.filter(HighwayAsset.asset_type.in_(["marking", "white_marking", "yellow_marking", "rpm", "delineator"]))
    assets = q.all()
    if not assets:
        raise HTTPException(404, f"No {standard} assets for {highway_id}")

    pdf = _compliance_pdf(highway_id, assets, standard)
    fname = f"{standard}_{highway_id}_{datetime.utcnow():%Y%m%d}.pdf"
    return Response(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/trend/pdf")
def trend_pdf(
    highway_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    _require_rl()
    q = db.query(Measurement).join(HighwayAsset)
    if highway_id:
        q = q.filter(HighwayAsset.highway_id == highway_id)
    measurements = q.order_by(Measurement.measured_at.desc()).limit(500).all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    flow = [
        Paragraph("Monthly Trend Analysis", styles["Heading1"]),
        Paragraph(
            f"Highway: {highway_id or 'All'} · Generated {datetime.utcnow():%Y-%m-%d %H:%M UTC}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    # Aggregate by month + source_layer
    from collections import defaultdict
    buckets: dict = defaultdict(list)
    for m in measurements:
        key = (m.measured_at.strftime("%Y-%m"), m.source_layer)
        buckets[key].append(m.rl_value)

    rows = [["Month", "Source", "N", "Mean R_L", "Min", "Max"]]
    for (month, src), vals in sorted(buckets.items()):
        rows.append([
            month, src, len(vals),
            round(sum(vals) / len(vals), 1),
            round(min(vals), 1),
            round(max(vals), 1),
        ])

    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    flow.append(t)
    doc.build(flow)

    return Response(
        buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="trend_analysis.pdf"'},
    )


# ── Excel: maintenance work order ───────────────────────────────────────────

@router.get("/work-order/xlsx")
def work_order_xlsx(
    highway_id: Optional[str] = Query(None),
    status: str = Query("pending"),
    db: Session = Depends(get_db),
):
    if not _HAS_XLSX:
        raise HTTPException(
            503, "openpyxl not installed. `pip install openpyxl` to enable Excel export."
        )

    q = db.query(MaintenanceOrder).join(HighwayAsset).filter(
        MaintenanceOrder.status == status
    )
    if highway_id:
        q = q.filter(HighwayAsset.highway_id == highway_id)
    orders = q.order_by(MaintenanceOrder.priority_score.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    header = [
        "Order ID", "Asset ID", "Highway", "Chainage km", "Asset Type",
        "Current R_L", "IRC Min R_L", "Priority", "Status",
        "Scheduled Date", "Notes", "Created",
    ]
    ws.append(header)
    header_fill = PatternFill("solid", fgColor="0F3057")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for o in orders:
        a = o.asset
        ws.append([
            o.id,
            a.id if a else None,
            a.highway_id if a else "",
            round(a.chainage_km, 2) if a else "",
            a.asset_type if a else "",
            round(a.current_rl, 1) if a and a.current_rl else "",
            round(a.irc_minimum_rl, 1) if a else "",
            o.priority_score,
            o.status,
            o.scheduled_date.strftime("%Y-%m-%d") if o.scheduled_date else "",
            (o.notes or "")[:200],
            o.created_at.strftime("%Y-%m-%d"),
        ])

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    fname = f"work_orders_{highway_id or 'all'}_{datetime.utcnow():%Y%m%d}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
